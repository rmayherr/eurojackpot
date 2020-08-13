import openpyxl
from configparser import ConfigParser
from downloader import get_from_config
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import os
import pandas

def write_to_xls_file(*args):
    """Write dataframe to spreadsheet"""

    try:
        print(f'Writing to {get_from_config("xls_output")}...', end='')
        if os.path.exists(get_from_config("xls_output")):
            os.remove(get_from_config("xls_output"))
        font_head = Font(name = "Arial",
                    size = "12",
                    bold = True,
                    color = 'FFFFFF'
                    )
        fill_head_green = PatternFill(start_color='09A014', end_color='09A014', \
                              fill_type='solid')
        file = get_from_config('xls_output')
        wb = Workbook()
        for pos, dataframe in enumerate(args):
            ws = wb.create_sheet("", pos)
            index_length = len(dataframe.index)
            for r in dataframe_to_rows(dataframe, index=True, header=True):
                ws.append(r)
            for cell in ws['A'] + ws[1]:
                cell.font = font_head
                cell.fill = fill_head_green
        wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    except Exception as e:
        print(f'Error occurred while writing to file! \n\t {e}')
    else:
        wb.save(filename = file)
        print(f'Done.')
